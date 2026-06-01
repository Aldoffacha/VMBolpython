from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, text, case
from datetime import date, datetime, timezone, timedelta
from typing import Optional
import io, csv

from app.database import get_db
from app.models.user import Auditoria, Administrador, Empleado, Cliente
from app.utils.dependencies import require_role

router = APIRouter(prefix="/admin/auditoria", tags=["admin-auditoria"])

BOL = timezone(timedelta(hours=-4))  # Bolivia UTC-4

def hoy_bolivia():
    return datetime.now(timezone.utc).astimezone(BOL).date()


def get_filtros(fecha_inicio, fecha_fin, tipo_usuario, accion, tabla):
    filtros = [
        Auditoria.fecha_auditoria >= fecha_inicio,
        Auditoria.fecha_auditoria < text(f"'{fecha_fin}'::date + INTERVAL '1 day'"),
    ]
    if tipo_usuario: filtros.append(Auditoria.tipo_usuario == tipo_usuario)
    if accion:       filtros.append(Auditoria.accion == accion)
    if tabla:        filtros.append(Auditoria.tabla_afectada == tabla)
    return filtros


def get_registros_query(db, filtros):
    return (
        db.query(
            Auditoria,
            func.coalesce(Administrador.nombre, Empleado.nombre, Cliente.nombre).label("nombre_usuario")
        )
        .outerjoin(Administrador, (Auditoria.id_usuario == Administrador.id_admin) & (Auditoria.tipo_usuario == "admin"))
        .outerjoin(Empleado,      (Auditoria.id_usuario == Empleado.id_empleado)   & (Auditoria.tipo_usuario == "empleado"))
        .outerjoin(Cliente,       (Auditoria.id_usuario == Cliente.id_cliente)     & (Auditoria.tipo_usuario == "cliente"))
        .filter(*filtros)
        .order_by(Auditoria.fecha_auditoria.desc())
    )


def defaults(fecha_inicio, fecha_fin):
    if not fecha_inicio:
        hoy = hoy_bolivia()
        fecha_inicio = hoy.replace(day=1)
    if not fecha_fin:
        fecha_fin = hoy_bolivia()
    return fecha_inicio, fecha_fin


# ─────────────────────────────────────────────
#  LISTAR
# ─────────────────────────────────────────────

@router.get("")
def get_auditoria(
    fecha_inicio:  date          = Query(default=None),
    fecha_fin:     date          = Query(default=None),
    tipo_usuario:  Optional[str] = Query(default=None),
    accion:        Optional[str] = Query(default=None),
    tabla:         Optional[str] = Query(default=None),
    pagina:        int           = Query(default=1, ge=1),
    por_pagina:    int           = Query(default=10),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    fecha_inicio, fecha_fin = defaults(fecha_inicio, fecha_fin)
    filtros = get_filtros(fecha_inicio, fecha_fin, tipo_usuario, accion, tabla)

    hoy_bol = hoy_bolivia()

    stats_q = db.query(
        func.count(Auditoria.id_auditoria).label("total_registros"),
        func.count(Auditoria.id_usuario.distinct()).label("usuarios_activos"),
        func.count(Auditoria.tabla_afectada.distinct()).label("tablas_afectadas"),
        func.sum(case((Auditoria.fecha_auditoria >= hoy_bol, 1), else_=0)).label("registros_hoy"),
    ).filter(*filtros).first()

    total = db.query(func.count(Auditoria.id_auditoria)).filter(*filtros).scalar() or 0
    total_paginas = -(-total // por_pagina)
    offset = (pagina - 1) * por_pagina

    registros_raw = get_registros_query(db, filtros).offset(offset).limit(por_pagina).all()

    return {
        "stats": {
            "total_registros":  stats_q.total_registros  or 0,
            "usuarios_activos": stats_q.usuarios_activos or 0,
            "tablas_afectadas": stats_q.tablas_afectadas or 0,
            "registros_hoy":    int(stats_q.registros_hoy or 0),
        },
        "total":         total,
        "total_paginas": total_paginas,
        "pagina_actual": pagina,
        "registros": [
            {
                "id_auditoria":    r.Auditoria.id_auditoria,
                "fecha":           r.Auditoria.fecha_auditoria.strftime("%d/%m/%Y %H:%M"),
                "nombre_usuario":  r.nombre_usuario or f"Usuario #{r.Auditoria.id_usuario}",
                "tipo_usuario":    r.Auditoria.tipo_usuario or "",
                "tabla_afectada":  r.Auditoria.tabla_afectada,
                "accion":          r.Auditoria.accion,
                "id_registro":     r.Auditoria.id_registro,
                "ip_address":      r.Auditoria.ip_address or "",
                "datos_anteriores": r.Auditoria.datos_anteriores or "",
                "datos_nuevos":     r.Auditoria.datos_nuevos or "",
            }
            for r in registros_raw
        ]
    }


# ─────────────────────────────────────────────
#  EXPORTAR CSV
# ─────────────────────────────────────────────

@router.get("/exportar/csv")
def exportar_csv(
    fecha_inicio:  date          = Query(default=None),
    fecha_fin:     date          = Query(default=None),
    tipo_usuario:  Optional[str] = Query(default=None),
    accion:        Optional[str] = Query(default=None),
    tabla:         Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    fecha_inicio, fecha_fin = defaults(fecha_inicio, fecha_fin)
    filtros   = get_filtros(fecha_inicio, fecha_fin, tipo_usuario, accion, tabla)
    registros = get_registros_query(db, filtros).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Fecha", "Usuario", "Tipo", "Tabla", "Acción", "ID Registro", "IP"])
    for r in registros:
        writer.writerow([
            r.Auditoria.id_auditoria,
            r.Auditoria.fecha_auditoria.strftime("%d/%m/%Y %H:%M"),
            r.nombre_usuario or f"Usuario #{r.Auditoria.id_usuario}",
            r.Auditoria.tipo_usuario,
            r.Auditoria.tabla_afectada,
            r.Auditoria.accion,
            r.Auditoria.id_registro,
            r.Auditoria.ip_address or "",
        ])

    output.seek(0)
    nombre = f"auditoria_{hoy_bolivia()}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )


# ─────────────────────────────────────────────
#  EXPORTAR EXCEL
# ─────────────────────────────────────────────

@router.get("/exportar/excel")
def exportar_excel(
    fecha_inicio:  date          = Query(default=None),
    fecha_fin:     date          = Query(default=None),
    tipo_usuario:  Optional[str] = Query(default=None),
    accion:        Optional[str] = Query(default=None),
    tabla:         Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fecha_inicio, fecha_fin = defaults(fecha_inicio, fecha_fin)
    filtros   = get_filtros(fecha_inicio, fecha_fin, tipo_usuario, accion, tabla)
    registros = get_registros_query(db, filtros).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"

    header_fill = PatternFill("solid", fgColor="1A0A0E")
    header_font = Font(bold=True, color="C1121F", size=11)
    border_side = Side(style="thin", color="9A031E")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    accion_colors = {"INSERT": "0D3320", "UPDATE": "3D2A00", "DELETE": "3D0A0A"}
    accion_fonts  = {"INSERT": "10B981", "UPDATE": "F59E0B", "DELETE": "EF4444"}

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Auditoría del Sistema — {fecha_inicio} a {fecha_fin}"
    ws["A1"].font      = Font(bold=True, color="C1121F", size=14)
    ws["A1"].fill      = PatternFill("solid", fgColor="0D0F12")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Fecha/Hora", "Usuario", "Tipo", "Tabla", "Acción", "ID Registro", "IP"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(registros, 3):
        accion_val = r.Auditoria.accion or ""
        row_fill   = PatternFill("solid", fgColor="1F2429" if row_idx % 2 == 0 else "161A1E")

        valores = [
            f"#{r.Auditoria.id_auditoria}",
            r.Auditoria.fecha_auditoria.strftime("%d/%m/%Y %H:%M"),
            r.nombre_usuario or f"Usuario #{r.Auditoria.id_usuario}",
            r.Auditoria.tipo_usuario or "",
            r.Auditoria.tabla_afectada or "",
            accion_val,
            f"#{r.Auditoria.id_registro}",
            r.Auditoria.ip_address or "—",
        ]

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center" if col in (1, 4, 5, 6, 7) else "left")
            if col == 6 and accion_val in accion_colors:
                cell.fill = PatternFill("solid", fgColor=accion_colors[accion_val])
                cell.font = Font(bold=True, color=accion_fonts[accion_val])
            else:
                cell.fill = row_fill
                cell.font = Font(color="D9D9D9")

    anchos = [8, 16, 22, 10, 12, 10, 12, 15]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"auditoria_{hoy_bolivia()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )


# ─────────────────────────────────────────────
#  EXPORTAR PDF
# ─────────────────────────────────────────────

@router.get("/exportar/pdf")
def exportar_pdf(
    fecha_inicio:  date          = Query(default=None),
    fecha_fin:     date          = Query(default=None),
    tipo_usuario:  Optional[str] = Query(default=None),
    accion:        Optional[str] = Query(default=None),
    tabla:         Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    fecha_inicio, fecha_fin = defaults(fecha_inicio, fecha_fin)
    filtros   = get_filtros(fecha_inicio, fecha_fin, tipo_usuario, accion, tabla)
    registros = get_registros_query(db, filtros).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles  = getSampleStyleSheet()
    rojo    = colors.HexColor("#9A031E")
    oscuro  = colors.HexColor("#121418")
    gris    = colors.HexColor("#D9D9D9")

    titulo_style = ParagraphStyle("titulo", parent=styles["Heading1"],
                                  textColor=rojo, fontSize=16, spaceAfter=4)
    sub_style    = ParagraphStyle("sub", parent=styles["Normal"],
                                  textColor=colors.HexColor("#A0A0A0"), fontSize=9, spaceAfter=12)

    elementos = [
        Paragraph("VMBol en Red — Auditoría del Sistema", titulo_style),
        Paragraph(f"Período: {fecha_inicio} al {fecha_fin}  |  Generado: {hoy_bolivia()}  |  Total: {len(registros)} registros", sub_style),
    ]

    cabecera = [["ID", "Fecha/Hora", "Usuario", "Tipo", "Tabla", "Acción", "ID Reg.", "IP"]]
    filas = cabecera + [
        [
            f"#{r.Auditoria.id_auditoria}",
            r.Auditoria.fecha_auditoria.strftime("%d/%m/%Y %H:%M"),
            (r.nombre_usuario or f"Usuario #{r.Auditoria.id_usuario}")[:22],
            r.Auditoria.tipo_usuario or "",
            r.Auditoria.tabla_afectada or "",
            r.Auditoria.accion or "",
            f"#{r.Auditoria.id_registro}",
            r.Auditoria.ip_address or "—",
        ]
        for r in registros
    ]

    col_widths = [1.5*cm, 3.5*cm, 5*cm, 2.2*cm, 2.8*cm, 2.2*cm, 2.2*cm, 3.5*cm]
    tabla      = Table(filas, colWidths=col_widths, repeatRows=1)

    accion_colors_pdf = {
        "INSERT": colors.HexColor("#10B981"),
        "UPDATE": colors.HexColor("#F59E0B"),
        "DELETE": colors.HexColor("#EF4444"),
    }

    style_cmds = [
        ("BACKGROUND",    (0, 0), (-1, 0),  oscuro),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  rojo),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  9),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#1F2429"), colors.HexColor("#161A1E")]),
        ("TEXTCOLOR",     (0, 1), (-1, -1), gris),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#9A031E")),
        ("ROWHEIGHT",     (0, 0), (-1, -1), 18),
    ]

    for i, r in enumerate(registros, 1):
        accion_val = r.Auditoria.accion or ""
        if accion_val in accion_colors_pdf:
            style_cmds.append(("TEXTCOLOR", (5, i), (5, i), accion_colors_pdf[accion_val]))
            style_cmds.append(("FONTNAME",  (5, i), (5, i), "Helvetica-Bold"))

    tabla.setStyle(TableStyle(style_cmds))
    elementos.append(tabla)

    doc.build(elementos)
    buf.seek(0)

    nombre = f"auditoria_{hoy_bolivia()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )