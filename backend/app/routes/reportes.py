from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, text
from datetime import date, datetime
from typing import Optional
import io, csv

from app.database import get_db
from app.models.user import Pedido, PedidoDetalle, Producto, Cliente
from app.utils.dependencies import require_role

router = APIRouter(prefix="/admin/reportes", tags=["admin-reportes"])


def build_filters(fecha_inicio, fecha_fin, producto_id, cliente_id):
    filtros = [Pedido.estado != "cancelado", Pedido.fecha.between(fecha_inicio, fecha_fin)]
    if producto_id:
        filtros.append(PedidoDetalle.id_producto == producto_id)
    if cliente_id:
        filtros.append(Pedido.id_cliente == cliente_id)
    return filtros


@router.get("")
def get_reportes(
    fecha_inicio: date  = Query(default=None),
    fecha_fin:    date  = Query(default=None),
    producto_id:  Optional[int] = Query(default=None),
    cliente_id:   Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    # Defaults: mes actual
    if not fecha_inicio:
        hoy = date.today()
        fecha_inicio = hoy.replace(day=1)
    if not fecha_fin:
        from calendar import monthrange
        hoy = date.today()
        fecha_fin = hoy.replace(day=monthrange(hoy.year, hoy.month)[1])

    filtros = build_filters(fecha_inicio, fecha_fin, producto_id, cliente_id)

    total_ventas  = db.query(func.sum(Pedido.total)).outerjoin(PedidoDetalle).filter(*filtros).scalar() or 0
    total_pedidos = db.query(func.count(Pedido.id_pedido.distinct())).outerjoin(PedidoDetalle).filter(*filtros).scalar() or 0
    total_clientes_nuevos = db.query(func.count(Cliente.id_cliente)).filter(
        Cliente.fecha_registro.between(fecha_inicio, fecha_fin)
    ).scalar() or 0
    ticket_promedio = float(total_ventas) / total_pedidos if total_pedidos else 0

    # ── Ventas mensuales (6 meses) ──
    ventas_mensuales_raw = db.execute(text("""
        SELECT EXTRACT(YEAR FROM fecha)::int AS anio,
               EXTRACT(MONTH FROM fecha)::int AS mes,
               SUM(total) AS total
        FROM pedidos
        WHERE fecha >= NOW() - INTERVAL '6 months' AND estado != 'cancelado'
        GROUP BY anio, mes ORDER BY anio, mes
    """)).fetchall()

    # ── Ventas semanales (8 semanas) ──
    ventas_semanales_raw = db.execute(text("""
        SELECT EXTRACT(YEAR FROM fecha)::int AS anio,
               EXTRACT(WEEK FROM fecha)::int AS semana,
               MIN(fecha)::date AS fecha_inicio,
               SUM(total) AS total
        FROM pedidos
        WHERE fecha >= NOW() - INTERVAL '8 weeks' AND estado != 'cancelado'
        GROUP BY anio, semana ORDER BY anio, semana DESC LIMIT 8
    """)).fetchall()

    # ── Distribución de estados ──
    estados_raw = (
        db.query(Pedido.estado, func.count(Pedido.id_pedido))
        .outerjoin(PedidoDetalle).filter(*filtros)
        .group_by(Pedido.estado).all()
    )

    # ── Top 10 productos más vendidos ──
    productos_raw = (
        db.query(
            Producto.nombre,
            func.sum(PedidoDetalle.cantidad).label("cantidad"),
            func.sum(PedidoDetalle.precio * PedidoDetalle.cantidad).label("total")
        )
        .join(PedidoDetalle, Producto.id_producto == PedidoDetalle.id_producto)
        .join(Pedido, PedidoDetalle.id_pedido == Pedido.id_pedido)
        .filter(*filtros)
        .group_by(Producto.id_producto, Producto.nombre)
        .order_by(func.sum(PedidoDetalle.cantidad).desc())
        .limit(10).all()
    )

    # ── Detalle ventas ──
    detalles_raw = (
        db.query(
            Pedido.id_pedido, Pedido.fecha, Pedido.estado,
            Cliente.nombre.label("cliente"),
            Producto.nombre.label("producto"),
            PedidoDetalle.cantidad, PedidoDetalle.precio,
        )
        .join(PedidoDetalle, Pedido.id_pedido == PedidoDetalle.id_pedido)
        .join(Cliente, Pedido.id_cliente == Cliente.id_cliente)
        .join(Producto, PedidoDetalle.id_producto == Producto.id_producto)
        .filter(*filtros)
        .order_by(Pedido.fecha.desc())
        .all()
    )

    # ── Listas para filtros ──
    lista_productos = db.query(Producto.id_producto, Producto.nombre).filter(Producto.estado == 1).order_by(Producto.nombre).all()
    lista_clientes  = db.query(Cliente.id_cliente, Cliente.nombre).filter(Cliente.estado == 1).order_by(Cliente.nombre).all()

    MESES = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    return {
        "filtros": {
            "fecha_inicio": str(fecha_inicio),
            "fecha_fin":    str(fecha_fin),
            "producto_id":  producto_id,
            "cliente_id":   cliente_id,
        },
        "stats": {
            "total_ventas":        float(total_ventas),
            "total_pedidos":       total_pedidos,
            "clientes_nuevos":     total_clientes_nuevos,
            "ticket_promedio":     round(ticket_promedio, 2),
        },
        "ventas_mensuales": [
            {"label": f"{MESES[r.mes]} {r.anio}", "total": float(r.total)} for r in ventas_mensuales_raw
        ],
        "ventas_semanales": [
            {"label": f"Sem {r.semana}", "total": float(r.total)} for r in reversed(ventas_semanales_raw)
        ],
        "distribucion_estados": [
            {"estado": est, "total": cnt} for est, cnt in estados_raw
        ],
        "productos_top": [
            {
                "nombre":   p.nombre,
                "cantidad": int(p.cantidad),
                "total":    float(p.total),
                "porcentaje": round(float(p.total) / float(total_ventas) * 100, 1) if total_ventas else 0,
            }
            for p in productos_raw
        ],
        "ventas_detalladas": [
            {
                "id_pedido": d.id_pedido,
                "fecha":     d.fecha.strftime("%d/%m/%Y %H:%M"),
                "cliente":   d.cliente,
                "producto":  d.producto,
                "cantidad":  d.cantidad,
                "precio":    float(d.precio),
                "total_linea": float(d.precio * d.cantidad),
                "estado":    d.estado,
            }
            for d in detalles_raw
        ],
        "lista_productos": [{"id": p.id_producto, "nombre": p.nombre} for p in lista_productos],
        "lista_clientes":  [{"id": c.id_cliente,  "nombre": c.nombre} for c in lista_clientes],
    }


# ── Exportar CSV ──
@router.get("/exportar/csv")
def exportar_csv(
    fecha_inicio: date  = Query(default=None),
    fecha_fin:    date  = Query(default=None),
    producto_id:  Optional[int] = Query(default=None),
    cliente_id:   Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    if not fecha_inicio:
        hoy = date.today(); fecha_inicio = hoy.replace(day=1)
    if not fecha_fin:
        from calendar import monthrange
        hoy = date.today(); fecha_fin = hoy.replace(day=monthrange(hoy.year, hoy.month)[1])

    filtros = build_filters(fecha_inicio, fecha_fin, producto_id, cliente_id)

    detalles = (
        db.query(
            Pedido.id_pedido, Pedido.fecha, Pedido.estado,
            Cliente.nombre.label("cliente"),
            Producto.nombre.label("producto"),
            PedidoDetalle.cantidad, PedidoDetalle.precio,
        )
        .join(PedidoDetalle, Pedido.id_pedido == PedidoDetalle.id_pedido)
        .join(Cliente, Pedido.id_cliente == Cliente.id_cliente)
        .join(Producto, PedidoDetalle.id_producto == Producto.id_producto)
        .filter(*filtros).order_by(Pedido.fecha.desc()).all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Pedido", "Fecha", "Cliente", "Producto", "Cantidad", "Precio", "Total", "Estado"])
    for d in detalles:
        writer.writerow([d.id_pedido, d.fecha.strftime("%d/%m/%Y %H:%M"), d.cliente, d.producto,
                         d.cantidad, f"${float(d.precio):.2f}", f"${float(d.precio * d.cantidad):.2f}", d.estado])

    output.seek(0)
    nombre = f"reporte_ventas_{date.today()}.csv"
    return StreamingResponse(io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )


# ── Exportar Excel ──
@router.get("/exportar/excel")
def exportar_excel(
    fecha_inicio: date  = Query(default=None),
    fecha_fin:    date  = Query(default=None),
    producto_id:  Optional[int] = Query(default=None),
    cliente_id:   Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not fecha_inicio:
        hoy = date.today(); fecha_inicio = hoy.replace(day=1)
    if not fecha_fin:
        from calendar import monthrange
        hoy = date.today(); fecha_fin = hoy.replace(day=monthrange(hoy.year, hoy.month)[1])

    filtros = build_filters(fecha_inicio, fecha_fin, producto_id, cliente_id)

    productos = (
        db.query(
            Producto.nombre,
            func.sum(PedidoDetalle.cantidad).label("cantidad"),
            func.sum(PedidoDetalle.precio * PedidoDetalle.cantidad).label("total")
        )
        .join(PedidoDetalle, Producto.id_producto == PedidoDetalle.id_producto)
        .join(Pedido, PedidoDetalle.id_pedido == Pedido.id_pedido)
        .filter(*filtros)
        .group_by(Producto.id_producto, Producto.nombre)
        .order_by(func.sum(PedidoDetalle.cantidad).desc())
        .all()
    )

    detalles = (
        db.query(
            Pedido.id_pedido, Pedido.fecha, Pedido.estado,
            Cliente.nombre.label("cliente"),
            Producto.nombre.label("producto"),
            PedidoDetalle.cantidad, PedidoDetalle.precio,
        )
        .join(PedidoDetalle, Pedido.id_pedido == PedidoDetalle.id_pedido)
        .join(Cliente, Pedido.id_cliente == Cliente.id_cliente)
        .join(Producto, PedidoDetalle.id_producto == Producto.id_producto)
        .filter(*filtros).order_by(Pedido.fecha.desc()).all()
    )

    total_ventas = db.query(func.sum(Pedido.total)).outerjoin(PedidoDetalle).filter(*filtros).scalar() or 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ventas"

    header_fill = PatternFill("solid", fgColor="1A0A0E")
    header_font = Font(bold=True, color="C1121F", size=11)
    border_side = Side(style="thin", color="9A031E")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Reporte de Ventas — {fecha_inicio} a {fecha_fin}"
    ws["A1"].font      = Font(bold=True, color="C1121F", size=14)
    ws["A1"].fill      = PatternFill("solid", fgColor="0D0F12")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Sección 1: Top Productos
    ws.merge_cells("A3:D3")
    ws["A3"] = "Top Productos Más Vendidos"
    ws["A3"].font = Font(bold=True, color="C1121F", size=12)
    ws["A3"].fill = PatternFill("solid", fgColor="0D0F12")

    headers_p = ["Producto", "Cantidad Vendida", "Total Vendido", "% del Total"]
    for col, h in enumerate(headers_p, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(productos, 5):
        pct = round(float(p.total) / float(total_ventas) * 100, 1) if total_ventas else 0
        row_fill = PatternFill("solid", fgColor="1F2429" if row_idx % 2 == 0 else "161A1E")
        for col, val in enumerate([p.nombre, int(p.cantidad), f"${float(p.total):.2f}", f"{pct}%"], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = thin_border
            cell.font      = Font(color="D9D9D9")
            cell.fill      = row_fill
            if col == 3:
                cell.font = Font(color="10B981", bold=True)

    # Sección 2: Detalle Ventas
    start_row = 5 + len(productos) + 2
    ws.merge_cells(f"A{start_row}:H{start_row}")
    ws.cell(row=start_row, column=1, value="Detalle de Ventas").font = Font(bold=True, color="C1121F", size=12)
    ws.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor="0D0F12")

    headers_d = ["Pedido", "Fecha", "Cliente", "Producto", "Cantidad", "Precio", "Total", "Estado"]
    for col, h in enumerate(headers_d, 1):
        cell = ws.cell(row=start_row + 1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, d in enumerate(detalles, start_row + 2):
        row_fill = PatternFill("solid", fgColor="1F2429" if row_idx % 2 == 0 else "161A1E")
        for col, val in enumerate([
            f"#{d.id_pedido}", d.fecha.strftime("%d/%m/%Y %H:%M"), d.cliente, d.producto,
            d.cantidad, f"${float(d.precio):.2f}", f"${float(d.precio * d.cantidad):.2f}", d.estado
        ], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = thin_border
            cell.font      = Font(color="D9D9D9")
            cell.fill      = row_fill
            if col == 7:
                cell.font = Font(color="10B981", bold=True)

    anchos = [10, 18, 22, 22, 10, 12, 14, 12]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"reporte_ventas_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )


# ── Exportar PDF ──
@router.get("/exportar/pdf")
def exportar_pdf(
    fecha_inicio: date  = Query(default=None),
    fecha_fin:    date  = Query(default=None),
    producto_id:  Optional[int] = Query(default=None),
    cliente_id:   Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("administrador"))
):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    if not fecha_inicio:
        hoy = date.today(); fecha_inicio = hoy.replace(day=1)
    if not fecha_fin:
        from calendar import monthrange
        hoy = date.today(); fecha_fin = hoy.replace(day=monthrange(hoy.year, hoy.month)[1])

    filtros = build_filters(fecha_inicio, fecha_fin, producto_id, cliente_id)

    productos = (
        db.query(
            Producto.nombre,
            func.sum(PedidoDetalle.cantidad).label("cantidad"),
            func.sum(PedidoDetalle.precio * PedidoDetalle.cantidad).label("total")
        )
        .join(PedidoDetalle, Producto.id_producto == PedidoDetalle.id_producto)
        .join(Pedido, PedidoDetalle.id_pedido == Pedido.id_pedido)
        .filter(*filtros)
        .group_by(Producto.id_producto, Producto.nombre)
        .order_by(func.sum(PedidoDetalle.cantidad).desc())
        .all()
    )

    detalles = (
        db.query(
            Pedido.id_pedido, Pedido.fecha, Pedido.estado,
            Cliente.nombre.label("cliente"),
            Producto.nombre.label("producto"),
            PedidoDetalle.cantidad, PedidoDetalle.precio,
        )
        .join(PedidoDetalle, Pedido.id_pedido == PedidoDetalle.id_pedido)
        .join(Cliente, Pedido.id_cliente == Cliente.id_cliente)
        .join(Producto, PedidoDetalle.id_producto == Producto.id_producto)
        .filter(*filtros).order_by(Pedido.fecha.desc()).all()
    )

    total_ventas = db.query(func.sum(Pedido.total)).outerjoin(PedidoDetalle).filter(*filtros).scalar() or 0
    total_pedidos = db.query(func.count(Pedido.id_pedido.distinct())).outerjoin(PedidoDetalle).filter(*filtros).scalar() or 0

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles  = getSampleStyleSheet()
    rojo    = colors.HexColor("#9A031E")
    oscuro  = colors.HexColor("#121418")
    gris    = colors.HexColor("#D9D9D9")

    titulo_style = ParagraphStyle("titulo", parent=styles["Heading1"],
                                  textColor=rojo, fontSize=16, spaceAfter=4)
    sub_style    = ParagraphStyle("sub", parent=styles["Normal"],
                                  textColor=colors.HexColor("#A0A0A0"), fontSize=9, spaceAfter=12)

    elementos = [
        Paragraph("VMBol en Red — Reporte de Ventas", titulo_style),
        Paragraph(f"Período: {fecha_inicio} al {fecha_fin}  |  Generado: {date.today()}  |  Total: {total_pedidos} pedidos, ${float(total_ventas):.2f} en ventas", sub_style),
    ]

    if productos:
        elementos.append(Paragraph("Top Productos Más Vendidos", ParagraphStyle("sec", parent=styles["Heading2"], textColor=rojo, fontSize=12, spaceAfter=6, spaceBefore=12)))
        cabecera_p = [["Producto", "Cantidad", "Total", "%"]]
        filas_p = cabecera_p + [
            [p.nombre[:30], str(int(p.cantidad)), f"${float(p.total):.2f}",
             f"{round(float(p.total) / float(total_ventas) * 100, 1) if total_ventas else 0}%"]
            for p in productos
        ]
        col_widths_p = [8*cm, 3*cm, 3.5*cm, 2*cm]
        tabla_p = Table(filas_p, colWidths=col_widths_p, repeatRows=1)
        tabla_p.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  oscuro),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  rojo),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  9),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#1F2429"), colors.HexColor("#161A1E")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), gris),
            ("FONTSIZE",      (0, 1), (-1, -1), 8),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#9A031E")),
            ("ROWHEIGHT",     (0, 0), (-1, -1), 18),
        ]))
        elementos.append(tabla_p)
        elementos.append(Spacer(1, 12))

    if detalles:
        elementos.append(Paragraph("Detalle de Ventas", ParagraphStyle("sec", parent=styles["Heading2"], textColor=rojo, fontSize=12, spaceAfter=6, spaceBefore=12)))
        cabecera_d = [["Pedido", "Fecha", "Cliente", "Producto", "Cant.", "Precio", "Total", "Estado"]]
        filas_d = cabecera_d + [
            [
                f"#{d.id_pedido}",
                d.fecha.strftime("%d/%m/%Y"),
                d.cliente[:18],
                d.producto[:20],
                str(d.cantidad),
                f"${float(d.precio):.2f}",
                f"${float(d.precio * d.cantidad):.2f}",
                d.estado,
            ]
            for d in detalles
        ]
        col_widths_d = [1.5*cm, 2.8*cm, 4*cm, 4.5*cm, 1.5*cm, 2.2*cm, 2.5*cm, 3*cm]
        tabla_d = Table(filas_d, colWidths=col_widths_d, repeatRows=1)
        tabla_d.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  oscuro),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  rojo),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  9),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#1F2429"), colors.HexColor("#161A1E")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), gris),
            ("FONTSIZE",      (0, 1), (-1, -1), 7),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#9A031E")),
            ("ROWHEIGHT",     (0, 0), (-1, -1), 16),
        ]))
        elementos.append(tabla_d)

    doc.build(elementos)
    buf.seek(0)

    nombre = f"reporte_ventas_{date.today()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )